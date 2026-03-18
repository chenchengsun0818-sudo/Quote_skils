import pandas as pd
import re
import os
import io
import json
import openpyxl
from openpyxl.styles import Alignment, Font, borders
from datetime import datetime

# --- 1. 基础路径配置 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DEFAULT_FILE = os.path.join(DATA_DIR, "default_price.xlsx")
DEFAULT_TEMPLATE = os.path.join(DATA_DIR, "default_template.xlsx")
CONFIG_FILE = os.path.join(DATA_DIR, "system_config.json")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")  # AI 专属输出文件夹

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)


# --- 2. 核心工具函数 (从原版提取) ---
def load_sys_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {"rate": 8.3, "points": 10.0}


def get_item_type(sku):
    s = str(sku).upper()
    if s.startswith('FC') or s.endswith('-DD') or 'CONTRACT' in s:
        return "Service"
    return "Hardware"


def clean_price(val):
    if pd.isna(val) or str(val).strip() in ["", "-", "n/a"]:
        return 0.0
    res = re.sub(r'[^\d.]', '', str(val))
    return float(res) if res else 0.0


def is_sku_pattern(val):
    s = str(val).strip()
    if s == 'nan' or s == "" or s.lower() == "sku" or len(s) < 4: return False
    if ' ' in s or any(c.islower() for c in s if c.isalpha()): return False
    return True


# --- 3. 数据加载与表格生成引擎 (从原版提取) ---
def load_and_parse_data(file):
    try:
        xl = pd.ExcelFile(file)
        blacklist = ["instruction", "notice", "summary", "contents", "cover sheet", "index", "general info",
                     "ordering guides", "changes", "dataset"]
        sheets = [s for s in xl.sheet_names if not any(w in s.lower() for w in blacklist)]

        all_data = []
        for s in sheets:
            df_raw = pd.read_excel(file, sheet_name=s, header=None)
            if df_raw.empty: continue
            col_idx = {'SKU': 1, 'DESC': 2, 'BASE': 3, 'P1': 4, 'P3': 6, 'P5': 8}
            header_row = -1
            for i in range(min(60, len(df_raw))):
                row_vals = [str(v).upper() for v in df_raw.iloc[i].values]
                if "SKU" in row_vals:
                    header_row = i
                    col_idx['SKU'] = row_vals.index("SKU")
                    for idx, v in enumerate(row_vals):
                        if "DESC" in v: col_idx['DESC'] = idx
                        if "PRICE" in v: col_idx['BASE'] = idx
                        if "1YR" in v: col_idx['P1'] = idx
                        if "3YR" in v: col_idx['P3'] = idx
                        if "5YR" in v: col_idx['P5'] = idx
                    break
            if header_row == -1: continue
            current_unit = "Other"
            for idx, row in df_raw.iterrows():
                if idx <= header_row: continue
                c0 = str(row[0]).strip()
                if c0.startswith(('Forti', 'FG-', 'FAP-')): current_unit = c0
                sku = str(row[col_idx['SKU']]).strip()
                if is_sku_pattern(sku):
                    all_data.append({
                        'UNIT': current_unit, 'SKU': sku, 'DESC': str(row[col_idx['DESC']]),
                        'BASE': row[col_idx['BASE']], 'P1': row[col_idx['P1']], 'P3': row[col_idx['P3']],
                        'P5': row[col_idx['P5']], 'Sheet': s, 'Type': get_item_type(sku)
                    })
        return pd.DataFrame(all_data)
    except Exception:
        return pd.DataFrame()


def fill_excel_template(template_file, data_df, meta_info):
    # 完全保留了原版 main.py 中你新增的 "=D*E" 和 "=SUM()" 等动态公式逻辑
    try:
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        base_font, bold_font, total_font = Font(name='仿宋', size=12), Font(name='仿宋', size=12, bold=True), Font(
            name='仿宋', size=13, bold=True)
        thin_border = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'),
                                     top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))
        pseudo_merge_left = borders.Border(left=borders.Side(style='thin'), top=borders.Side(style='thin'),
                                           bottom=borders.Side(style='thin'))
        pseudo_merge_mid = borders.Border(top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))
        pseudo_merge_right = borders.Border(right=borders.Side(style='thin'), top=borders.Side(style='thin'),
                                            bottom=borders.Side(style='thin'))
        align_center, align_left, align_right, align_wrap_left = Alignment(horizontal='center',
                                                                           vertical='center'), Alignment(
            horizontal='left', vertical='center'), Alignment(horizontal='right', vertical='center'), Alignment(
            horizontal='left', vertical='center', wrap_text=True)

        replacements = {
            "{{Customer}}": meta_info['customer'], "{{Project}}": meta_info['project'], "{{Agent}}": meta_info['agent'],
            "{{Date}}": datetime.now().strftime("%Y-%m-%d"), "{{Sales_Name}}": meta_info['s_name'],
            "{{Sales_Phone}}": meta_info['s_phone'],
            "{{Sales_Email}}": meta_info['s_email'], "{{Total_Amount}}": meta_info['total_str']
        }

        table_start_row = -1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "{{Table_Start}}" in cell.value:
                        table_start_row = cell.row
                        cell.value = ""
                        continue
                    for key, val in replacements.items():
                        if key in cell.value: cell.value = cell.value.replace(key, str(val))

        if table_start_row > 0 and not data_df.empty:
            units = data_df['UNIT'].unique()
            total_rows_to_insert = len(data_df) + (len(units) * 2) + 1
            ws.insert_rows(table_start_row + 1, amount=total_rows_to_insert - 1)
            current_row = table_start_row
            subtotal_rows = []

            for unit in units:
                unit_df = data_df[data_df['UNIT'] == unit]
                for col in range(1, 7):
                    c = ws.cell(row=current_row, column=col)
                    c.border = pseudo_merge_left if col == 1 else (pseudo_merge_right if col == 6 else pseudo_merge_mid)

                hw_rows = unit_df[unit_df['Type'] == 'Hardware']
                group_title = str(hw_rows.iloc[0]['SKU']) if not hw_rows.empty else str(unit)
                c1_header = ws.cell(row=current_row, column=1, value=group_title)
                c1_header.alignment, c1_header.font = align_left, base_font
                current_row += 1
                item_idx = 1
                start_item_row = current_row

                for _, row in unit_df.iterrows():
                    c1 = ws.cell(row=current_row, column=1, value=item_idx)
                    c2 = ws.cell(row=current_row, column=2, value=row['SKU'])
                    c3 = ws.cell(row=current_row, column=3, value=row['Desc'])
                    c4 = ws.cell(row=current_row, column=4, value=row['Unit(¥)'])
                    c5 = ws.cell(row=current_row, column=5, value=row['Qty'])
                    c6 = ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}")

                    for cx in (c1, c2, c3, c4, c5, c6): cx.font, cx.border = base_font, thin_border
                    c1.alignment, c4.alignment, c5.alignment, c6.alignment = align_center, align_center, align_center, align_center
                    c2.alignment, c3.alignment = align_left, align_wrap_left
                    c4.number_format = c6.number_format = '¥#,##0'

                    item_idx += 1;
                    current_row += 1

                end_item_row = current_row - 1
                for col in range(1, 6):
                    c = ws.cell(row=current_row, column=col)
                    c.border = pseudo_merge_left if col == 1 else (pseudo_merge_right if col == 5 else pseudo_merge_mid)

                c5_sub = ws.cell(row=current_row, column=5, value="小计")
                c5_sub.alignment, c5_sub.font = align_right, bold_font
                c6_sub = ws.cell(row=current_row, column=6, value=f"=SUM(F{start_item_row}:F{end_item_row})")
                c6_sub.border, c6_sub.alignment, c6_sub.font, c6_sub.number_format = thin_border, align_center, bold_font, '¥#,##0'
                subtotal_rows.append(current_row)
                current_row += 1

            for col in range(1, 6):
                c = ws.cell(row=current_row, column=col)
                c.border = pseudo_merge_left if col == 1 else (pseudo_merge_right if col == 5 else pseudo_merge_mid)
            c5_tot = ws.cell(row=current_row, column=5, value="总计")
            c5_tot.alignment, c5_tot.font = align_right, total_font

            formula_str = "=" + "+".join([f"F{r}" for r in subtotal_rows]) if subtotal_rows else "=0"
            c6_tot = ws.cell(row=current_row, column=6, value=formula_str)
            c6_tot.border, c6_tot.alignment, c6_tot.font, c6_tot.number_format = thin_border, align_center, total_font, '¥#,##0'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        print(f"Error in Excel generation: {e}")
        return None


# --- 4. 🤖 AI 主调用入口 ---
def generate_quote_skill(customer: str, project: str, sku_list: list, discount: float = 35.0):
    """
    提供给 OpenClaw 的核心执行接口
    """
    full_data = load_and_parse_data(DEFAULT_FILE)
    if full_data.empty:
        return {"status": "error", "message": "未能加载价格数据库 default_price.xlsx"}

    # 读取默认配置汇率和过单点
    sys_cfg = load_sys_config()
    ex_rate = sys_cfg.get("rate", 8.3)
    dist_points = sys_cfg.get("points", 10.0)

    calc_res = []
    not_found = []
    grand_total = 0.0

    # 模拟 UI 的核心算价与匹配逻辑
    for sku in sku_list:
        sku = sku.strip().upper()
        match = full_data[full_data['SKU'] == sku]
        duration_val = 'Fixed'

        if match.empty and re.search(r'-(12|36|60)$', sku):
            base_sku = re.sub(r'-(12|36|60)$', '-DD', sku)
            match = full_data[full_data['SKU'] == base_sku]
            if sku.endswith('-12'):
                duration_val = '1 Yr'
            elif sku.endswith('-36'):
                duration_val = '3 Yr'
            elif sku.endswith('-60'):
                duration_val = '5 Yr'

        if not match.empty:
            item = match.iloc[0].to_dict()
            if item.get('Type') == 'Service' and duration_val == 'Fixed':
                duration_val = '3 Yr'

            # 根据 Duration 找基础价
            if item['Type'] == 'Hardware':
                base_usd = clean_price(item.get('BASE', 0))
            else:
                p_map = {"1 Yr": "P1", "3 Yr": "P3", "5 Yr": "P5", "Fixed": "BASE"}
                key = p_map.get(duration_val, "BASE")
                base_usd = clean_price(item.get(key, 0)) or clean_price(item.get('BASE', 0))

            # 根据你在 main.py 中的核心公式计算
            sell_cny = base_usd * (1 - discount / 100) * ex_rate / (1 - dist_points / 100)

            f_sku = item['SKU']
            if "-DD" in f_sku:
                sf = {"1 Yr": "-12", "3 Yr": "-36", "5 Yr": "-60"}.get(duration_val, "-36")
                f_sku = f_sku.replace("-DD", sf)

            qty = 1  # AI 默认提单数量为 1
            total_cny = sell_cny * qty
            grand_total += total_cny

            calc_res.append({
                "UNIT": item.get('UNIT', 'Other'), "Type": item.get('Type', 'Other'),
                "SKU": f_sku, "Desc": item['DESC'], "ListPrice($)": base_usd,
                "Discount": discount, "Unit(¥)": sell_cny, "Qty": qty, "Total(¥)": total_cny
            })
        else:
            not_found.append(sku)

    if not calc_res:
        return {"status": "error", "message": f"未找到任何有效型号。未识别列表: {', '.join(not_found)}"}

    df_calc = pd.DataFrame(calc_res)
    meta = {
        'customer': customer, 'project': project, 'agent': 'AI 直客',
        's_name': "智能助理", 's_phone': "", 's_email': "",
        'total_str': f"¥ {grand_total:,.2f}"
    }

    excel_data = fill_excel_template(DEFAULT_TEMPLATE, df_calc, meta)
    if not excel_data:
        return {"status": "error", "message": "Excel 模版渲染失败，请检查 default_template.xlsx"}

    # 保存实体文件供 OpenClaw 回传
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_cust = re.sub(r'[\\/*?:"<>|]', "", customer)
    file_name = f"Quote_{safe_cust}_{timestamp}.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)

    with open(file_path, 'wb') as f:
        f.write(excel_data)

    msg = f"✅ 核价完成，总计 ¥ {grand_total:,.2f}。"
    if not_found: msg += f"\n⚠️ 以下型号未在库中找到: {', '.join(not_found)}"

    return {
        "status": "success",
        "message": msg,
        "total_amount": grand_total,
        "file_path": file_path
    }